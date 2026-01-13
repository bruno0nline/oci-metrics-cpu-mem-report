import os
import csv
import json
from datetime import datetime

import oci
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ================= CONFIG =================
HOME = os.path.expanduser("~")
CSV_PATH = os.path.join(HOME, "Relatorio_Instancias_Tags_OCI.csv")
XLSX_PATH = os.path.join(HOME, "Relatorio_Instancias_Tags_OCI.xlsx")
# =========================================

cfg = oci.config.from_file()
tenancy_id = cfg["tenancy"]

identity = oci.identity.IdentityClient(cfg)


def get_regions():
    return [r.region_name for r in identity.list_region_subscriptions(tenancy_id).data]


def get_compartments():
    comps = oci.pagination.list_call_get_all_results(
        identity.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True
    ).data
    root = identity.get_compartment(tenancy_id).data
    return [c for c in comps if c.lifecycle_state == "ACTIVE"] + [root]


def main():
    regions = get_regions()
    compartments = get_compartments()
    rows = []

    print("\n📦 Coletando Inventário de Instâncias + Tags\n")

    for region in regions:
        print(f"🟢 Região: {region}")

        cfg_r = dict(cfg)
        cfg_r["region"] = region

        compute = oci.core.ComputeClient(cfg_r)

        for comp in compartments:
            try:
                instances = oci.pagination.list_call_get_all_results(
                    compute.list_instances,
                    compartment_id=comp.id
                ).data
            except oci.exceptions.ServiceError:
                continue

            if not instances:
                continue

            print(f"  📁 {comp.name} | Instâncias: {len(instances)}")

            for inst in instances:
                rows.append({
                    "region": region,
                    "compartment": comp.name,
                    "instance_name": inst.display_name,
                    "instance_state": inst.lifecycle_state,
                    "shape": inst.shape,
                    "ocpus": getattr(inst.shape_config, "ocpus", None),
                    "memory_gb": getattr(inst.shape_config, "memory_in_gbs", None),
                    "freeform_tags": json.dumps(inst.freeform_tags or {}),
                    "defined_tags": json.dumps(inst.defined_tags or {})
                })

    # ================= CSV =================
    headers = list(rows[0].keys())

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    # ================= XLSX =================
    wb = Workbook()
    ws = wb.active
    ws.title = "INSTANCIAS_TAGS"
    ws.append(headers)

    fill_running = PatternFill("solid", fgColor="C6EFCE")
    fill_stopped = PatternFill("solid", fgColor="FFC7CE")

    col_state = headers.index("instance_state") + 1

    for r in rows:
        ws.append([r[h] for h in headers])
        row = ws.max_row
        if r["instance_state"] == "RUNNING":
            ws.cell(row=row, column=col_state).fill = fill_running
        else:
            ws.cell(row=row, column=col_state).fill = fill_stopped

    wb.save(XLSX_PATH)

    print("\n✅ Relatórios gerados:")
    print(f"➡ CSV : {CSV_PATH}")
    print(f"➡ XLSX: {XLSX_PATH}")


if __name__ == "__main__":
    main()
