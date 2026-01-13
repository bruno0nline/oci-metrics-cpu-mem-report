import csv
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ================= CONFIG =================
HOME = os.path.expanduser("~")

INPUT_CSV = os.path.join(HOME, "Relatorio_Instancias_Tags_OCI.csv")
OUT_CSV = os.path.join(HOME, "Relatorio_FinOps_StartStop.csv")
OUT_XLSX = os.path.join(HOME, "Relatorio_FinOps_StartStop.xlsx")
# =========================================


def extract_tag(tags, *keys):
    if not tags:
        return ""
    for k in keys:
        if k in tags:
            return str(tags.get(k))
    return ""


def parse_json(value):
    try:
        return json.loads(value) if value else {}
    except Exception:
        return {}


def main():
    rows_out = []

    with open(INPUT_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for r in reader:
            freeform = parse_json(r.get("freeform_tags") or r.get("all_freeform_tags"))

            env = extract_tag(freeform, "Environment", "Env")
            autostop = extract_tag(freeform, "AutoStop", "autostop", "Schedule")
            owner = extract_tag(freeform, "Owner", "Responsavel")
            cost = extract_tag(freeform, "CostCenter", "CentroCusto")

            has_autostop = "YES" if autostop.lower() in ("true", "yes", "1") else "NO"

            status = "OK"
            if r["instance_state"] == "RUNNING" and has_autostop == "NO":
                status = "RISK"

            rows_out.append({
                "region": r["region"],
                "compartment": r["compartment"],
                "instance_name": r["instance_name"],
                "state": r["instance_state"],
                "shape": r["shape"],
                "ocpus": r["ocpus"],
                "memory_gb": r["memory_gb"],
                "environment": env,
                "autostop": has_autostop,
                "owner": owner,
                "cost_center": cost,
                "finops_status": status
            })

    # ================= CSV =================
    headers = list(rows_out[0].keys())
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows_out)

    # ================= XLSX =================
    wb = Workbook()
    ws = wb.active
    ws.title = "FINOPS"

    ws.append(headers)

    fill_ok = PatternFill("solid", fgColor="C6EFCE")
    fill_risk = PatternFill("solid", fgColor="FFC7CE")

    col_status = headers.index("finops_status") + 1

    for r in rows_out:
        ws.append([r[h] for h in headers])
        row = ws.max_row
        ws.cell(row=row, column=col_status).fill = (
            fill_risk if r["finops_status"] == "RISK" else fill_ok
        )

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(OUT_XLSX)

    print("\n✅ Relatórios FinOps gerados:")
    print(f"➡ CSV : {OUT_CSV}")
    print(f"➡ XLSX: {OUT_XLSX}")


if __name__ == "__main__":
    main()
