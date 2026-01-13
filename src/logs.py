import os
import csv
from datetime import datetime, timezone

import oci
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ================= CONFIG =================
HOME = os.path.expanduser("~")
CSV_PATH = os.path.join(HOME, "Relatorio_FinOps_OCI_Logs.csv")
XLSX_PATH = os.path.join(HOME, "Relatorio_FinOps_OCI_Logs.xlsx")

NOISY_SERVICES = [
    "flowlogs",
    "vcn",
    "audit",
    "oke",
    "loadbalancer"
]
# =========================================

cfg = oci.config.from_file()
tenancy_id = cfg["tenancy"]

identity = oci.identity.IdentityClient(cfg)


def finops_recommendation(log_type, lifecycle, source_service):
    source_service = (source_service or "").lower()

    if lifecycle != "ACTIVE":
        return "REMOVE"

    if log_type == "CUSTOM":
        return "REVIEW"

    if any(s in source_service for s in NOISY_SERVICES):
        return "REVIEW"

    return "KEEP"


def main():
    rows = []

    regions = [r.region_name for r in identity.list_region_subscriptions(tenancy_id).data]

    compartments = oci.pagination.list_call_get_all_results(
        identity.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True
    ).data

    root = identity.get_compartment(tenancy_id).data
    compartments.append(root)

    for region in regions:
        print(f"\n🌎 Região: {region}")
        cfg_r = dict(cfg)
        cfg_r["region"] = region

        logging_client = oci.logging.LoggingManagementClient(cfg_r)

        for comp in compartments:
            try:
                log_groups = oci.pagination.list_call_get_all_results(
                    logging_client.list_log_groups,
                    compartment_id=comp.id
                ).data
            except Exception:
                continue

            for lg in log_groups:
                try:
                    logs = oci.pagination.list_call_get_all_results(
                        logging_client.list_logs,
                        log_group_id=lg.id
                    ).data
                except Exception:
                    continue

                for log in logs:
                    source = log.configuration.source if log.configuration else None

                    source_service = getattr(source, "service", None)
                    source_resource = getattr(source, "resource", None)

                    recommendation = finops_recommendation(
                        log.log_type,
                        log.lifecycle_state,
                        source_service
                    )

                    rows.append({
                        "region": region,
                        "compartment": comp.name,
                        "log_group": lg.display_name,
                        "log_name": log.display_name,
                        "log_type": log.log_type,
                        "lifecycle_state": log.lifecycle_state,
                        "source_service": source_service,
                        "source_resource": source_resource,
                        "time_created": log.time_created.strftime("%Y-%m-%d"),
                        "finops_recommendation": recommendation
                    })

    # ================= CSV =================
    headers = list(rows[0].keys())

    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    # ================= EXCEL =================
    wb = Workbook()
    ws = wb.active
    ws.title = "OCI Logs FinOps"
    ws.append(headers)

    fill_remove = PatternFill("solid", fgColor="FFC7CE")
    fill_review = PatternFill("solid", fgColor="FFEB9C")
    fill_keep = PatternFill("solid", fgColor="C6EFCE")

    rec_col = headers.index("finops_recommendation") + 1

    for r in rows:
        ws.append([r[h] for h in headers])
        row = ws.max_row
        rec = r["finops_recommendation"]

        if rec == "REMOVE":
            ws.cell(row=row, column=rec_col).fill = fill_remove
        elif rec == "REVIEW":
            ws.cell(row=row, column=rec_col).fill = fill_review
        else:
            ws.cell(row=row, column=rec_col).fill = fill_keep

    wb.save(XLSX_PATH)

    print("\n✅ Relatórios gerados:")
    print(CSV_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
