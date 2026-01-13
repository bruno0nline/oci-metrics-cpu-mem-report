import os
import csv
import time
from datetime import datetime, timedelta, timezone

import oci
from oci.monitoring.models import SummarizeMetricsDataDetails
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ================= CONFIG =================
DAYS = int(os.getenv("METRICS_DAYS", "30"))
INTERVAL = "5m"

CPU_LOW = 5
MEM_LOW = 40
CPU_HIGH = 80
MEM_HIGH = 85

HOME = os.path.expanduser("~")
CSV_PATH = os.path.join(HOME, f"Relatorio_FinOps_CPU_MEM_{DAYS}d.csv")
XLSX_PATH = os.path.join(HOME, f"Relatorio_FinOps_CPU_MEM_{DAYS}d.xlsx")
# =========================================

cfg = oci.config.from_file()
tenancy_id = cfg["tenancy"]
identity = oci.identity.IdentityClient(cfg)


def mean_p95(values):
    if not values:
        return None, None
    values = sorted(values)
    mean = sum(values) / len(values)
    p95 = values[int(len(values) * 0.95) - 1]
    return mean, p95


def get_metric(monitoring, compartment_id, instance_id, metric, start, end):
    query = f'{metric}[{INTERVAL}]{{resourceId = "{instance_id}"}}.mean()'
    details = SummarizeMetricsDataDetails(
        namespace="oci_computeagent",
        query=query,
        start_time=start,
        end_time=end
    )
    resp = monitoring.summarize_metrics_data(
        compartment_id=compartment_id,
        summarize_metrics_data_details=details
    )
    if not resp.data or not resp.data[0].aggregated_datapoints:
        return None, None
    values = [d.value for d in resp.data[0].aggregated_datapoints if d.value is not None]
    return mean_p95(values)


def finops_recommendation(cpu_mean, cpu_p95, mem_mean, mem_p95):
    cpu_mean = cpu_mean or 0
    mem_mean = mem_mean or 0
    cpu_p95 = cpu_p95 or 0
    mem_p95 = mem_p95 or 0

    if cpu_mean < CPU_LOW and mem_mean < MEM_LOW:
        return "DOWNSIZE-STRONG"
    if cpu_p95 > CPU_HIGH or mem_p95 > MEM_HIGH:
        return "UPSCALE"
    return "KEEP"


def main():
    start = datetime.now(timezone.utc) - timedelta(days=DAYS)
    end = datetime.now(timezone.utc)

    rows = []

    regions = [r.region_name for r in identity.list_region_subscriptions(tenancy_id).data]
    compartments = oci.pagination.list_call_get_all_results(
        identity.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True
    ).data

    root = identity.get_compartment(tenancy_id).data
    compartments.append(root)

    for region in regions:
        cfg_r = dict(cfg)
        cfg_r["region"] = region

        compute = oci.core.ComputeClient(cfg_r)
        monitoring = oci.monitoring.MonitoringClient(cfg_r)

        for comp in compartments:
            try:
                instances = oci.pagination.list_call_get_all_results(
                    compute.list_instances,
                    compartment_id=comp.id
                ).data
            except Exception:
                continue

            for inst in instances:
                if inst.lifecycle_state != "RUNNING":
                    continue

                cpu_mean, cpu_p95 = get_metric(
                    monitoring, comp.id, inst.id, "CpuUtilization", start, end
                )
                mem_mean, mem_p95 = get_metric(
                    monitoring, comp.id, inst.id, "MemoryUtilization", start, end
                )

                rows.append({
                    "region": region,
                    "compartment": comp.name,
                    "instance_name": inst.display_name,
                    "shape": inst.shape,
                    "ocpus": getattr(inst.shape_config, "ocpus", None),
                    "memory_gb": getattr(inst.shape_config, "memory_in_gbs", None),
                    "cpu_mean_percent": cpu_mean,
                    "cpu_p95_percent": cpu_p95,
                    "mem_mean_percent": mem_mean,
                    "mem_p95_percent": mem_p95,
                    "finops_recommendation": finops_recommendation(
                        cpu_mean, cpu_p95, mem_mean, mem_p95
                    )
                })

    # CSV
    headers = list(rows[0].keys())
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "FinOps"
    ws.append(headers)

    fill_down = PatternFill("solid", fgColor="FFC7CE")
    fill_keep = PatternFill("solid", fgColor="C6EFCE")
    fill_up = PatternFill("solid", fgColor="FFEB9C")

    rec_col = headers.index("finops_recommendation") + 1

    for r in rows:
        ws.append([r[h] for h in headers])
        row = ws.max_row
        rec = r["finops_recommendation"]
        if rec == "DOWNSIZE-STRONG":
            ws.cell(row=row, column=rec_col).fill = fill_down
        elif rec == "UPSCALE":
            ws.cell(row=row, column=rec_col).fill = fill_up
        else:
            ws.cell(row=row, column=rec_col).fill = fill_keep

    wb.save(XLSX_PATH)

    print("✅ Relatórios gerados:")
    print(CSV_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
