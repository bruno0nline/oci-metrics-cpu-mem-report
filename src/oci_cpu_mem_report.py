import os
import time
import csv
from datetime import datetime, timedelta, timezone

import oci
from oci.monitoring.models import SummarizeMetricsDataDetails
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ================= CONFIG =================
DAYS = int(os.getenv("METRICS_DAYS", "30"))
INTERVAL = "5m"

CPU_LOW = 5
CPU_MED = 15
CPU_HIGH = 80

MEM_LOW = 40
MEM_HIGH = 85

MAX_RETRIES = 3
RETRY_SLEEP = 3

HOME = os.path.expanduser("~")
CSV_PATH = os.path.join(HOME, f"Relatorio_CPU_MEM_{DAYS}d.csv")
XLSX_PATH = os.path.join(HOME, f"Relatorio_CPU_MEM_{DAYS}d.xlsx")
# =========================================

cfg = oci.config.from_file()
tenancy_id = cfg["tenancy"]
identity = oci.identity.IdentityClient(cfg)


def get_regions():
    return [r.region_name for r in identity.list_region_subscriptions(tenancy_id).data]


def get_compartments():
    comps = oci.pagination.list_call_get_all_results(
        identity.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True
    ).data
    root = identity.get_compartment(tenancy_id).data
    return [c for c in comps if c.lifecycle_state == "ACTIVE"] + [root]


def summarize_with_retry(monitoring, compartment_id, details):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return monitoring.summarize_metrics_data(
                compartment_id=compartment_id,
                summarize_metrics_data_details=details
            )
        except oci.exceptions.ServiceError as e:
            if e.status == 429 and attempt < MAX_RETRIES:
                time.sleep(RETRY_SLEEP)
                continue
            raise


def mean_p95(values):
    if not values:
        return None, None
    values = sorted(values)
    mean = sum(values) / len(values)
    p95 = values[int(len(values) * 0.95) - 1]
    return mean, p95


def get_metric(monitoring, compartment_id, instance_id, metric, start, end):
    query = f'{metric}[{INTERVAL}]{{resourceId = "{instance_id}"}}.mean()'
    details = SummarizeMetricsDataDetails(
        namespace="oci_computeagent",
        query=query,
        start_time=start,
        end_time=end,
    )
    resp = summarize_with_retry(monitoring, compartment_id, details)
    if not resp.data or not resp.data[0].aggregated_datapoints:
        return None, None
    values = [d.value for d in resp.data[0].aggregated_datapoints if d.value is not None]
    return mean_p95(values)


def finops(cpu_mean, cpu_p95, mem_mean, mem_p95):
    cpu_mean = cpu_mean or 0
    mem_mean = mem_mean or 0
    cpu_p95 = cpu_p95 or 0
    mem_p95 = mem_p95 or 0

    if cpu_mean < CPU_LOW and mem_mean < MEM_LOW:
        return "DOWNSIZE-STRONG"
    if cpu_mean < CPU_MED and mem_mean < 60:
        return "DOWNSIZE"
    if mem_mean < MEM_LOW:
        return "DOWNSIZE-MEM"
    if cpu_p95 > CPU_HIGH or mem_p95 > MEM_HIGH:
        return "UPSCALE"
    return "KEEP"


def main():
    regions = get_regions()
    compartments = get_compartments()

    start = datetime.now(timezone.utc) - timedelta(days=DAYS)
    end = datetime.now(timezone.utc)

    rows = []

    print(f"\n📊 Coletando CPU/Memória ({DAYS} dias)\n")

    for region in regions:
        print(f"🟢 Região: {region}")
        cfg_r = dict(cfg)
        cfg_r["region"] = region

        compute = oci.core.ComputeClient(cfg_r)
        monitoring = oci.monitoring.MonitoringClient(cfg_r)

        for comp in compartments:
            try:
                instances = oci.pagination.list_call_get_all_results(
                    compute.list_instances,
                    compartment_id=comp.id
                ).data
            except oci.exceptions.ServiceError:
                print(f"⚠️ Sem acesso ao compartment {comp.name}")
                continue

            running = [i for i in instances if i.lifecycle_state == "RUNNING"]
            if not running:
                continue

            print(f"  📁 {comp.name} | RUNNING: {len(running)}")

            for inst in running:
                cpu_mean, cpu_p95 = get_metric(
                    monitoring, comp.id, inst.id, "CpuUtilization", start, end
                )
                mem_mean, mem_p95 = get_metric(
                    monitoring, comp.id, inst.id, "MemoryUtilization", start, end
                )

                rows.append({
                    "region": region,
                    "compartment": comp.name,
                    "instance_name": inst.display_name,
                    "shape": inst.shape,
                    "ocpus": getattr(inst.shape_config, "ocpus", None),
                    "memory_gb": getattr(inst.shape_config, "memory_in_gbs", None),
                    "cpu_mean_percent": cpu_mean,
                    "cpu_p95_percent": cpu_p95,
                    "mem_mean_percent": mem_mean,
                    "mem_p95_percent": mem_p95,
                    "finops_recommendation": finops(cpu_mean, cpu_p95, mem_mean, mem_p95),
                })

    headers = list(rows[0].keys())

    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "CPU_MEM"

    ws.append(headers)

    fill_keep = PatternFill("solid", fgColor="C6EFCE")
    fill_down = PatternFill("solid", fgColor="FFC7CE")
    fill_up = PatternFill("solid", fgColor="FFEB9C")

    rec_col = headers.index("finops_recommendation") + 1

    for r in rows:
        ws.append([r[h] for h in headers])
        row = ws.max_row
        rec = r["finops_recommendation"]
        if rec.startswith("DOWNSIZE"):
            ws.cell(row=row, column=rec_col).fill = fill_down
        elif rec == "UPSCALE":
            ws.cell(row=row, column=rec_col).fill = fill_up
        else:
            ws.cell(row=row, column=rec_col).fill = fill_keep

    wb.save(XLSX_PATH)

    print("\n✅ Relatórios gerados:")
    print(f"➡ CSV : {CSV_PATH}")
    print(f"➡ XLSX: {XLSX_PATH}")


if __name__ == "__main__":
    main()
