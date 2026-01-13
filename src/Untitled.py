#!/usr/bin/env python3
"""
Read-only FinOps recommender para OCI (CPU / Memória)

- 100% read-only: não faz nenhuma operação que altere recursos.
- Coleta métricas (média/p95) e sugere ações:
    - DOWNSIZE / DOWNSIZE-STRONG / DOWNSIZE-MEM / KEEP / UPSCALE
- Sugestão de OCPUs alvo calculada a partir do OCPUs atual e do p95 vs threshold.
- Salva CSV e XLSX com recomendações.

Uso exemplo:
    python src/oci_finops_readonly_recommend.py --days 30 --outdir ~/finops_reports
"""
from __future__ import annotations

import csv
import logging
import math
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import oci
from oci.monitoring.models import SummarizeMetricsDataDetails
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ---------- Defaults / thresholds ----------
DAYS = int(os.getenv("METRICS_DAYS", "30"))
INTERVAL = os.getenv("METRICS_INTERVAL", "5m")

CPU_LOW = int(os.getenv("CPU_LOW", "5"))
CPU_MED = int(os.getenv("CPU_MED", "15"))
CPU_HIGH = int(os.getenv("CPU_HIGH", "80"))

MEM_LOW = int(os.getenv("MEM_LOW", "40"))
MEM_HIGH = int(os.getenv("MEM_HIGH", "85"))

MAX_RETRIES = int(os.getenv("METRICS_MAX_RETRIES", "3"))
RETRY_SLEEP = float(os.getenv("METRICS_RETRY_SLEEP", "1.5"))
# ------------------------------------------

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("oci-finops-readonly")


def mean_p95(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    if not values:
        return None, None
    vals = sorted(values)
    mean = sum(vals) / len(vals)
    # p95 via rank interpolation
    n = len(vals)
    rank = 0.95 * (n - 1)
    lo = int(math.floor(rank))
    hi = int(math.ceil(rank))
    if lo == hi:
        p95 = vals[int(rank)]
    else:
        frac = rank - lo
        p95 = vals[lo] * (1 - frac) + vals[hi] * frac
    return mean, p95


def summarize_with_retry(monitoring: oci.monitoring.MonitoringClient, compartment_id: str, details: SummarizeMetricsDataDetails):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return monitoring.summarize_metrics_data(compartment_id=compartment_id, summarize_metrics_data_details=details)
        except oci.exceptions.ServiceError as e:
            status = getattr(e, "status", None)
            logger.debug("summarize_metrics_data ServiceError status=%s attempt=%d", status, attempt)
            if status == 429 and attempt < MAX_RETRIES:
                import time
                time.sleep(RETRY_SLEEP * attempt)
                continue
            raise


def get_metric(monitoring: oci.monitoring.MonitoringClient, compartment_id: str, instance_id: str, metric: str, start: datetime, end: datetime) -> Tuple[Optional[float], Optional[float]]:
    query = f'{metric}[{INTERVAL}]{{resourceId = "{instance_id}"}}.mean()'
    details = SummarizeMetricsDataDetails(namespace="oci_computeagent", query=query, start_time=start, end_time=end)
    resp = summarize_with_retry(monitoring, compartment_id, details)
    if not resp.data or not resp.data[0].aggregated_datapoints:
        return None, None
    values = [float(d.value) for d in resp.data[0].aggregated_datapoints if d.value is not None]
    return mean_p95(values)


def parse_baseline(instance: Any) -> Tuple[str, str, str]:
    shape_cfg = getattr(instance, "shape_config", None)
    baseline = getattr(shape_cfg, "baseline_ocpu_utilization", None)
    if not baseline:
        return "NO", "Desativada", ""
    mapping = {"BASELINE_1_8": "12.5%", "BASELINE_1_2": "50%", "BASELINE_1_1": "100%"}
    return "YES", mapping.get(baseline, baseline), baseline


def finops_recommendation(cpu_mean: Optional[float], cpu_p95: Optional[float], mem_mean: Optional[float], mem_p95: Optional[float]) -> str:
    cm = cpu_mean or 0
    mm = mem_mean or 0
    cp95 = cpu_p95 or 0
    mp95 = mem_p95 or 0
    if cm < CPU_LOW and mm < MEM_LOW:
        return "DOWNSIZE-STRONG"
    if cm < CPU_MED and mm < 60:
        return "DOWNSIZE"
    if mm < MEM_LOW:
        return "DOWNSIZE-MEM"
    if cp95 > CPU_HIGH or mp95 > MEM_HIGH:
        return "UPSCALE"
    return "KEEP"


def suggest_ocpus_increase(current_ocpus: Optional[float], cpu_p95: Optional[float]) -> Optional[int]:
    """
    Sugere OCPUs alvo baseado em p95 vs threshold CPU_HIGH.
    Fórmula simples: target = ceil(current * (cpu_p95 / CPU_HIGH))
    Se current_ocpus não disponível, retorna None.
    """
    if current_ocpus is None or cpu_p95 is None:
        return None
    if cpu_p95 <= CPU_HIGH:
        return None
    factor = cpu_p95 / CPU_HIGH
    target = int(math.ceil(current_ocpus * factor))
    # garantir pelo menos +1 se target == current
    if target <= current_ocpus:
        target = int(current_ocpus) + 1
    return target


def build_row(region: str, comp_name: str, inst: Any, inst_full: Any, cpu_mean: Optional[float], cpu_p95: Optional[float], mem_mean: Optional[float], mem_p95: Optional[float]) -> Dict[str, Any]:
    ocpus = getattr(getattr(inst_full, "shape_config", None), "ocpus", None)
    mem_gb = getattr(getattr(inst_full, "shape_config", None), "memory_in_gbs", None)
    burst, baseline, baseline_raw = parse_baseline(inst_full)

    rec = finops_recommendation(cpu_mean, cpu_p95, mem_mean, mem_p95)
    suggested_ocpus = None
    suggested_action = None
    if rec == "UPSCALE":
        suggested_ocpus = suggest_ocpus_increase(ocpus, cpu_p95)
        if suggested_ocpus:
            suggested_action = f"Increase OCPUs -> {suggested_ocpus} (current: {ocpus})"
        else:
            suggested_action = "Consider larger shape (no ocpus info)"

    elif rec.startswith("DOWNSIZE"):
        suggested_action = "Consider smaller shape / fewer OCPUs"

    else:
        suggested_action = "No action"

    def fmt(v):
        return None if v is None else round(v, 2)

    return {
        "region": region,
        "compartment": comp_name,
        "instance_name": inst.display_name,
        "instance_ocid": inst.id,
        "shape": inst.shape,
        "ocpus": ocpus,
        "memory_gb": mem_gb,
        "burstable_enabled": burst,
        "baseline_percent": baseline,
        "cpu_mean_percent": fmt(cpu_mean),
        "cpu_p95_percent": fmt(cpu_p95),
        "mem_mean_percent": fmt(mem_mean),
        "mem_p95_percent": fmt(mem_p95),
        "finops_recommendation": rec,
        "suggested_action": suggested_action,
        "suggested_ocpus": suggested_ocpus
    }


def save_csv(rows: List[Dict[str, Any]], path: Path):
    headers = list(rows[0].keys())
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    logger.info("CSV salvo em: %s", path)


def save_xlsx(rows: List[Dict[str, Any]], path: Path):
    headers = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    ws.append(headers)
    fill_up = PatternFill("solid", fgColor="FFEB9C")
    fill_down = PatternFill("solid", fgColor="FFC7CE")
    fill_keep = PatternFill("solid", fgColor="C6EFCE")
    rec_col = headers.index("finops_recommendation") + 1
    for r in rows:
        ws.append([r.get(h) for h in headers])
        row = ws.max_row
        rec = r.get("finops_recommendation", "")
        if isinstance(rec, str) and rec.startswith("DOWNSIZE"):
            ws.cell(row=row, column=rec_col).fill = fill_down
        elif rec == "UPSCALE":
            ws.cell(row=row, column=rec_col).fill = fill_up
        else:
            ws.cell(row=row, column=rec_col).fill = fill_keep
    wb.save(str(path))
    logger.info("XLSX salvo em: %s", path)


def main(days: int = DAYS, outdir: Optional[str] = None):
    outdir = Path(outdir or Path.home()).expanduser()
    csv_path = outdir / f"finops_recommendations_{days}d.csv"
    xlsx_path = outdir / f"finops_recommendations_{days}d.xlsx"

    cfg = oci.config.from_file()
    tenancy_id = cfg["tenancy"]
    identity = oci.identity.IdentityClient(cfg)

    regions = [r.region_name for r in identity.list_region_subscriptions(tenancy_id).data]
    compartments = oci.pagination.list_call_get_all_results(identity.list_compartments, tenancy_id, compartment_id_in_subtree=True).data
    root = identity.get_compartment(tenancy_id).data
    compartments = [c for c in compartments if c.lifecycle_state == "ACTIVE"] + [root]

    start = datetime.now(timezone.utc) - timedelta(days=days)
    end = datetime.now(timezone.utc)

    rows: List[Dict[str, Any]] = []
    logger.info("Coletando métricas (read-only) para %d dias em %d regiões", days, len(regions))

    for region in regions:
        logger.info("Região: %s", region)
        cfg_r = dict(cfg)
        cfg_r["region"] = region
        compute = oci.core.ComputeClient(cfg_r)
        monitoring = oci.monitoring.MonitoringClient(cfg_r)

        for comp in compartments:
            try:
                instances = oci.pagination.list_call_get_all_results(compute.list_instances, compartment_id=comp.id).data
            except Exception:
                continue
            running = [i for i in instances if i.lifecycle_state == "RUNNING"]
            if not running:
                continue
            logger.info("  %s RUNNING=%d", comp.name, len(running))
            for inst in running:
                try:
                    inst_full = compute.get_instance(inst.id).data  # leitura apenas
                except Exception:
                    continue
                cpu_mean, cpu_p95 = get_metric(monitoring, comp.id, inst.id, "CpuUtilization", start, end)
                mem_mean, mem_p95 = get_metric(monitoring, comp.id, inst.id, "MemoryUtilization", start, end)
                row = build_row(region, comp.name, inst, inst_full, cpu_mean, cpu_p95, mem_mean, mem_p95)
                rows.append(row)

    if not rows:
        logger.warning("Nenhuma instância encontrada ou sem dados.")
        return

    rows.sort(key=lambda r: (r["region"], r["compartment"], r["instance_name"]))
    save_csv(rows, csv_path)
    save_xlsx(rows, xlsx_path)
    logger.info("Recomendações geradas: %s , %s", csv_path, xlsx_path)
    # imprimir amostra de recomendações importantes
    ups = [r for r in rows if r["finops_recommendation"] == "UPSCALE"]
    if ups:
        logger.info("Instances sugeridas para UPSCALE (exibindo até 10):")
        for r in ups[:10]:
            logger.info("- %s (%s) - %s", r["instance_name"], r["instance_ocid"], r["suggested_action"])


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--days", type=int, default=DAYS)
    p.add_argument("--outdir", type=str, default=str(Path.home()))
    args = p.parse_args()
    main(days=args.days, outdir=args.outdir)